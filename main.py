# Librairies de manipulation de répertoires
import os
import os.path
from pathlib import Path
# Librairie permet la construction de l'interface graphique
from tkinter import *
from tkinter import ttk
from tkinter.ttk import Progressbar
from tkinter.filedialog import *
# Librairie de manipulation de tableau 
import pandas as pd
# Librairie de manipulation de REGEX
import re
# Librairie de manipulation des fichiers tableur
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
# Librairie pour la sauvegarde des valeurs à remplacer
from functools import partial
# Librairie pour lancer un nouveau thread dans le but d'affichier une barre de progression
import threading
# Librairie pour récupérer des valeurs depuis l'horloge (pour nom du fichier final)
import time

# Modules personnels
from thesaurus_tools.insee import get_insee # permet de récupérer l'INSEE, liste de corresondance dans le module
from thesaurus_tools.thesaurusconverter import ColCorrespMerimee, ColCorrespPalissy # fonctions de calcul de correspondance entre l'ancien format du thésaurus et le nouveu
from thesaurus_tools.approximate import calculate_proximity # fonction de calcul de proximité entre les termes (utilise la librairie RapidFuzz)
from thesaurus_tools.interface import Window, Frame, Button, SheetPopup, ProvenancePopup, get_coor # classes sur mesure pour l'interface
from thesaurus_tools.stringtools import adapt_to_col # fonction définissant des règles particulières à chacun des colonnes pour la comparaison des termes 


my_green = "spring green2" # couleur utilisée pour les boutons
str_to_be_compiled = "AFFE PAFF VISI GESTION BIBL BIBLIO ARCHIV SOURCE" # contient les valeurs préremplies des champs à compiler dans GESTION
str_to_ignore = "INSEE" # valeurs préremplies des colonnes à ignorer

# Fonction de conversion de la valeur de la cellule en une liste de valeurs lorsqu'elle sont séparées par ";"
def atomize(cell_str):
    if ";" in str(cell_str):
        cell_list = [term.strip() for term in cell_str.split(";")]
        return cell_list
    else:
        return [cell_str]

# Fonction qui appelle le popup permettant de sélectionner la feuille du fichier excel à traiter
def give_provenance(my_list, term, col_name):
    popup = ProvenancePopup(window, my_list, sheet, term, col_name)

# Fonction permettant d'inscrire la valeur à remplacer dans le dictionnaire des futurs changements si le bouton est pressé
def replace_term(term, new_term, col_first_cell, my_list, voc_button):
    if not isinstance(new_term, str):
        new_term = new_term.get()

    # Si le bouton était déjà vert, alors un pression le desactive et supprimer la valeur à remplacer
    if voc_button["background"] == my_green:
        voc_button.configure(bg=button_base_color)
        my_dict[col_first_cell][term].pop("replace")
    # Sinon, le bouton passe vert et la valeur est inscrite dans le dictionnaire des futurs changements
    else:
        voc_button.configure(bg=my_green)
        my_dict[col_first_cell][term]["replace"] = new_term

# Fonction d'affichage du dictionnaires contenant les erreurs
def add_logs(my_dict):
    for col_first_cell in my_dict:
        text.insert("end", f"\n\n{'='*50} {col_first_cell.value}")

        for term in my_dict[col_first_cell]:
            text.insert("end", f"\n\n\n\"{term}\"\t", "bold")
            term_button = Button(text)
            text.window_create("end", window=term_button)
            term_button.configure(text="Voir provenance", command=partial(give_provenance, my_dict[col_first_cell][term]["from"], term, col_first_cell.value))

            for new_term in my_dict[col_first_cell][term]["voc"]:
                global voc_button

                if new_term != "$":
                    text.insert("end", f"\n\n\t")
                    voc_button = Button(text)
                    text.window_create("end", window=voc_button)
                    voc_button.configure(text="Remplacer par :", command=partial(replace_term, term, new_term, col_first_cell, my_dict[col_first_cell][term]["from"], voc_button))
                    text.insert("end", f"\t\"{new_term}\"")
                
                else:
                    text.insert("end", f"\n\n\t")
                    voc_button = Button(text)
                    text.window_create("end", window=voc_button)
                    text.insert("end", f"  ")
                    user_voc = Entry(text, bg="gray95", width=40)
                    text.window_create("end", window=user_voc)
                    voc_button.configure(text="Remplacer par :", command=partial(replace_term, term, user_voc, col_first_cell, my_dict[col_first_cell][term]["from"], voc_button))                


    total = 0
    for col_first_cell in my_dict:
        total += len(my_dict[col_first_cell])
    text.insert("1.0", f"Cas détectés : {len(my_dict)}. Occurrences totales : {total}.")

# Fonctions permettant la mise en place de la barre de progression
def schedule_check(t):
    window.after(250, check_if_done, t)
def check_if_done(t):
    if not t.is_alive():
        progress_bar.stop()
    else:
        schedule_check(t)
def begin_search():
    text.delete('1.0', END)
    progress_bar.start()
    t = threading.Thread(target=checker)
    t.start()
    schedule_check(t)

# Fonction 
def checker():   
    global my_dict
    my_dict = {}
    for col in sheet.iter_cols(): # pour chaque colonne dans le feuille
        col_name_cindoc = col[0].value # récupération du nom de la colonne
        if col_name_cindoc and col_name_cindoc not in ign_ent.get().split(" "): # si entête colonne n'est pas vide ET colonne pas dans les colonnes à ignorer
            # Récupération de la correspondance en fonction du type de thésaurus
            if option_var_thes.get() == "Mérimée":
                corresp = ColCorrespMerimee(col_name_cindoc, thesaurus)
            elif option_var_thes.get() == "Palissy":
                corresp = ColCorrespPalissy(col_name_cindoc, thesaurus)
            refs = corresp.make_voc_list()
        
        if refs: # si refs (liste des termes du thésaurus de la colonne) existe (c à d ColCorrespXXXXXXX.make_voc_list a trouvé que la colonne avait une liste de voc thésaurus)
            my_dict[col[0]] = {}
            for cell in col[1:]: # pour chaque cellule
                if cell.value: # si la cellule n'est pas vide
                    for term in atomize(cell.value): # pour chacun des termes de la cellule
                        term = str(term) # conversion en string si jamais une des valeurs ne contient que des chiffres
                        if term and not term[0] == "=": # si le terme existe bien et ne commence pas par "=" (donc n'est pas une fonction excel)
                            term_part = adapt_to_col(term, col_name_cindoc) # application des règles propres à la colonne en question pour éviter les faux positifs
                            if cell and isinstance(term_part, str) and term_part not in refs: # si le terme n'est pas dans la liste de référence
                                # Constitution d'une entrée du dictionnaire pour le cas en question qui pose problème
                                if term in my_dict[col[0]]: # si ce terme a déjà été rencontré
                                    my_dict[col[0]][term]["from"].append(cell) # on ajoute le fait que ce cas existe aussi pour la cellule actuellement traitée
                                else: # si le terme n'est pas dans le dictionnaire : on l'ajoute
                                    my_dict[col[0]][term] = {} # nouveau dictionnaire
                                    my_dict[col[0]][term]["from"] = [cell] # on ajoute la cellule de provencance
                                    proxs = calculate_proximity(term, refs, option_var_alg.get(), int(num_ent.get())) # calculs des termes proches
                                    my_dict[col[0]][term]["voc"] = proxs # ajouts des termes proches (c à d des alternatives qui seront proposées à l'utilisateur)
                            elif cell and isinstance(term_part, list): # si le terme est une liste de plusieurs termes : même travail que plus haut mais pour chacun des termes
                                for part in term_part: 
                                    if part not in refs:
                                        if term in my_dict[col[0]]:
                                            my_dict[col[0]][term]["from"].append(cell)
                                        else:
                                            my_dict[col[0]][term] = {}
                                            my_dict[col[0]][term]["from"] = [cell]                     
                                            proxs = calculate_proximity(term, refs, option_var_alg.get(), int(num_ent.get()))
                                            my_dict[col[0]][term]["voc"] = proxs
    add_logs(my_dict) # une fois le dictionnaire constitué on affiche les résultats dans l'interface
    search_button.configure(bg=my_green) # le bouton de recherche qui avait appelé ces traitements passe vert

# Fonction d'ouverture du fichier excel : une variable pour l'excel, une pour la feuille, une pour le nom de fichier qui servira à nommer la sauvegarde
def open_excel():
    global excel
    global sheet
    global file_name
    file_path = askopenfilename(title="Sélection du fichier Excel", filetypes=[("Fichiers Excel", '.xlsx .xlsm .xlxt .xltm')])
    file_name = Path(file_path).stem
    excel = load_workbook(filename=file_path)
    if len(excel.sheetnames) > 1:
        sheets = [sheet for sheet in excel.sheetnames]
        sheet = excel[SheetPopup(window, sheets).get_name()]
    else:
        sheet = excel.active
    excel_button.configure(text=f"{os.path.basename(file_path)} - {sheet.title}", bg=my_green)

# Fonction d'ouverture du thésaurus, stockage dans la variable gloable thesaurus sous forme de dataframe pandas
def open_thesaurus():
    global thesaurus
    file_path = askopenfilename(title="Sélection du thésaurus", filetypes=[("Fichier texte", '*.csv')])
    thesaurus = pd.read_csv(file_path)
    thesaurus = thesaurus[["typeThesaurus", "libelle"]].groupby("typeThesaurus")["libelle"].apply(list)
    thesaurus = thesaurus.to_dict()
    thesau_button.configure(text=os.path.basename(file_path), bg=my_green)

# Fonction de compilation des colonnes choisies dans la colonne GESTION
def merge_col():
    try:
        col_gestion_old = get_column_letter(get_coor(sheet=sheet, col_name="GESTION").column)
        sheet[f"{col_gestion_old}1"] = "GESTION_old"
    except:
        pass

    col_gestion = get_column_letter(sheet.max_column + 1)
    sheet[f"{col_gestion}1"] = "GESTION_new"
    list_to_be_compiled = user_txt.get().replace("GESTION", "GESTION_old").split(" ")
    
    for col in list_to_be_compiled:
        try:
            cell_col = get_column_letter(get_coor(sheet=sheet, col_name=col).column)
            for i in range(2, sheet.max_row + 1):
                if sheet[f'{cell_col}{i}'].value:
                    if sheet[f"{col_gestion}{i}"].value:
                        sheet[f"{col_gestion}{i}"] = sheet[f"{col_gestion}{i}"].value + f"{{{col}={sheet[f'{cell_col}{i}'].value}}}"
                    else:
                        sheet[f"{col_gestion}{i}"] = f"{{{col}={sheet[f'{cell_col}{i}'].value}}}"
        except:
            pass

# Fonction de complétion de l'INSEE
def fill_insee():
    col_insee = get_column_letter(get_coor(sheet=sheet, col_name="INSEE").column) # récupération de la colonne INSEE
    col_com = get_column_letter(get_coor(sheet=sheet, col_name="COM").column) # récupération de la colonne COM

    # Pour toutes les cellules de la colonne
    for i in range(2, sheet.max_row + 1):
        # Si pas d'INSEE et que la COMMUNE est pourtant présente
        if not sheet[f"{col_insee}{i}"].value and sheet[f"{col_com}{i}"].value:
            sheet[f"{col_insee}{i}"] = get_insee(sheet[f"{col_com}{i}"].value) # inscription de l'INSEE

# Fonction de remplacement d'un terme au milieu de autres
def replace(cell_value, term, new_term):
    if ";" in cell_value:
        cell_values = cell_value.split(";")
        for i in range(len(cell_values)):
            cell_values[i] = cell_values[i].strip()
            if term in cell_values[i]:
                cell_values[i] = new_term
        cell_value = " ; ".join(cell_values)
    else:
        cell_value = new_term
    return cell_value

# Fonction de sauvegarde du fichier
def save_to():
    fill_insee() # complétion de l'INSEE si non présent
    if check_var.get() == 1: # si la compilation est sélectionné
        merge_col() # compilation dans la colonne GESTION

    # Boucle d'application des modifications inscrites dans le dictionnaire des futurs changement
    for col_first_cell in my_dict:
        for term in my_dict[col_first_cell]:
            if "replace" in my_dict[col_first_cell][term]:
                new_term = my_dict[col_first_cell][term]["replace"]
                for cell in my_dict[col_first_cell][term]["from"]:
                    cell_value = cell.value
                    sheet[cell.coordinate] = replace(cell_value, term, new_term)

    # Constitution du nom du fichier : inscription de la date / si date déjà dans le nom, remplacement de la date
    if re.sub(r"\_[0-9]{2}-[0-9]{2}-[0-9]{4}_[0-9]{2}h[0-9]{2}", "", file_name):
        name = re.sub(r"\_[0-9]{2}-[0-9]{2}-[0-9]{4}_[0-9]{2}h[0-9]{2}", "", file_name)
    else:
        name = file_name
    named_tuple = time.localtime() # get struct_time
    time_string = time.strftime("%d-%m-%Y_%Hh%M", named_tuple)
    new_file_name = f"{name}_{time_string}"

    # Sauvegarde effective du fichier excel modifié
    file_path = asksaveasfilename(parent=window, title="Sauvegarder sous...", initialfile=new_file_name, defaultextension=".xlsx", filetypes=[("Fichiers Excel", '.xlsx .xlsm .xlxt .xltm')])
    excel.save(file_path)
    save_button.configure(bg=my_green)

# Partie interface
if __name__ == "__main__":
    window = Window()
    my_pady = 10

    # Frame 00
    frame00 = Frame(window, row=0, column=0)
    wraplength = frame00.winfo_width()
    readme = """Ce programme permet de comparer les termes des colonnes/champs des exports Cindoc avec un thésaurus de référence."""
    readme_label = Label(frame00, text=readme, wraplength=530, justify=LEFT)
    readme_label.pack(expand=True, padx=8)

    # Frame 01
    frame01 = Frame(window, row=1, column=0)
    # Texte thésaurus
    thesau_text = Label(frame01, text="Sélection du fichier CSV ")
    thesau_text.pack(side=LEFT, padx=(8,0))
    # Bouton selection répertoire thésaurus
    thesau_button = Button(frame01, text="Ouvrir", command=open_thesaurus)
    thesau_button.pack(side=LEFT)
    # Option thésaurus
    option_text = Label(frame01, text=" du thésaurus ")
    option_text.pack(side=LEFT)
    option_var_thes = StringVar(frame01)
    option_thes = ttk.OptionMenu(frame01, option_var_thes, "Mérimée", *("Mérimée", "Palissy"))
    option_thes.pack(side=LEFT)
    
    # Frame 02
    frame02 = Frame(window, row=2, column=0)
    # Texte TSV
    tsv_text = Label(frame02, text="Sélection du fichier d'export Cindoc :")
    tsv_text.pack(side=LEFT, padx=(8,0))
    # Bouton selection répertoire thésaurus
    excel_button = Button(frame02, text="Ouvrir", command=open_excel)
    excel_button.pack(side=LEFT)

    # Frame 03
    frame03 = Frame(window, row=3, column=0)
    # Champ des colonnes à ne pas tenir compte
    ign_txt = Label(frame03, text="Colonnes à ignorer :")
    ign_txt.pack(side=LEFT, padx=(8,0))
    ign_ent = Entry(frame03, width=66)
    ign_ent.insert(END, str_to_ignore)
    ign_ent.pack(side=LEFT)

    # Frame 04
    frame04 = Frame(window, row=4, column=0)
    # Option thésaurus
    algo_text = Label(frame04, text="Algorithme : ")
    algo_text.pack(side=LEFT, padx=(8,0))
    option_var_alg = StringVar(frame04)
    option_alg = ttk.OptionMenu(frame04, option_var_alg, "Partial Ratio", *("Partial Ratio", "Simple Ratio", "Token Sort Ratio"))
    option_alg.pack(side=LEFT)
    num_text = Label(frame04, text="Nombre de termes proposés : ")
    num_text.pack(side=LEFT, padx=(8,0))
    num_ent = Entry(frame04, width=3)
    num_ent.insert(END, "5")
    num_ent.pack(side=LEFT)

    # Frame 05
    frame05 = Frame(window, row=5, column=0)
    # Barre de progression
    progress_bar = Progressbar(frame05, orient=HORIZONTAL, mode="determinate", length=400)
    progress_bar.pack(expand=True, pady=(5,0))
    # Bouton de recherche
    search_button = Button(frame05, text="Rechercher", command=begin_search)
    search_button.pack(expand=True)

    # Frame 06
    frame06 = Frame(window, row=6, column=0)
    # Champ des colonnes à compiler dans GESTION
    check_var = IntVar()
    check = Checkbutton(frame06, text='Compiler dans GESTION_new :', variable=check_var, onvalue=1, offvalue=0)
    check.pack(side=LEFT, padx=(8,0))
    user_txt = Entry(frame06, width=53)
    user_txt.insert(END, str_to_be_compiled)
    user_txt.pack(side=LEFT)

    # Frame 07
    frame07 = Frame(window, row=7, column=0)
    # Bouton de sauvegarde
    save_button = Button(frame07, text="Sauvegarder", command=save_to)
    save_button.pack(expand=True)
    
    # Frame 10
    frame10 = Frame(window, row=0, column=1, rowspan=8)
    frame10.columnconfigure(0, weight=1)
    # Champ de résultat
    text = Text(frame10, font="TkTextFont 9")
    text.tag_configure("bold", font="TkTextFont 9 bold")
    text.grid(row=0, column=0, sticky=EW)
    sb = Scrollbar(frame10, command=text.yview)
    sb.grid(row=0, column=0, sticky="NSE")
    text.configure(yscrollcommand=sb.set)

    window.mainloop()