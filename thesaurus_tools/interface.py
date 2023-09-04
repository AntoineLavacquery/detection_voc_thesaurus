from tkinter import *
from tkinter import ttk
# from tkinter.ttk import *
from openpyxl.utils import get_column_letter
import base64
import os
from thesaurus_tools.data import *

icon_data = base64.b64decode(icon)
temp_file = "icon.ico"
icon_file = open(temp_file, "wb")
icon_file.write(icon_data)
icon_file.close()

button_base_color = "LightBlue3"

def get_coor(sheet, col_name):
    for col in sheet.iter_cols():
        if col[0].value == col_name:
            return col[0]

class Window(Tk):
    def __init__(self):
        Tk.__init__(self)
        self.wm_title("Détecteur de vocabulaire thésaurus")
        self.configure(bg="grey88")
        self.geometry('1130x385')
        self.wm_iconbitmap(temp_file)
        os.remove(temp_file)

class Frame(Frame):
    def __init__(self, *args, **kwargs):
        super().__init__(*args)
        self.grid(**kwargs, padx=(10, 0), pady=(10, 0), sticky=NSEW)
        self.columnconfigure(0, weight=1)
        self.columnconfigure(1, weight=1)

class Button(Button):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, bg=button_base_color, activebackground='gold', relief="flat", **kwargs)

class SheetPopup(Toplevel):
    def __init__(self, parent, sheets):
        Toplevel.__init__(self, parent)
        self.x = parent.winfo_x()
        self.y = parent.winfo_y()
        self.geometry("+%d+%d" %(self.x+380, self.y+200))
        self.lift()

        self.var = StringVar()
        self.var.set("Sélectionner une feuille")

        self.label = Label(self, text="Merci de sélectionner la feuille à corriger :")
        self.option = ttk.OptionMenu(self, self.var, sheets[0], *sheets) # sheets[0]
        self.ok_button = Button(self, text="Ok", command=self.on_ok)

        params = {"side":"top", "fill":"x", "padx": 10, "pady": 5}
        self.label.pack(**params)
        self.option.pack(**params)
        self.ok_button.pack(**params)

    def on_ok(self):
        self.destroy()

    def get_name(self):
        self.wm_deiconify()
        self.wait_window()
        return self.var.get()

class ProvenancePopup(Toplevel):
    def __init__(self, parent, cell_list, sheet, term, col_name):
        Toplevel.__init__(self, parent)
        self.x = parent.winfo_x()
        self.y = parent.winfo_y()
        self.geometry("+%d+%d" %(self.x+260, self.y+130))
        self.cell_list = cell_list
        self.sheet = sheet
        self.term = term
        self.col_name = col_name
        self.lift()

        text = Text(self, font="TkTextFont 9")
        text.tag_configure("bold", font="TkTextFont 9 bold")
        text.pack(side="left", padx=10, pady=10)
        sb = Scrollbar(self, command=text.yview)
        sb.pack(side="right", fill="y")
        text.configure(yscrollcommand=sb.set)

        for cell in cell_list:
            REF = get_coor(sheet, "REF")
            ref = sheet[f"{get_column_letter(REF.column)}{cell.row}"].value
            text.insert("end", f"{self.col_name}\n{ref}\ncellule : {cell.coordinate}\n")

            cell_value = cell.value
            cell_list = cell_value.split(term)
            if cell_list[0] == "":
                text.insert("end", f"\"{term}", "bold")
                text.insert("end", f"{cell_list[-1]}\"")
            elif cell_list[-1] == "":
                text.insert("end", f"\"{cell_list[0]}")
                text.insert("end", f"{term}", "bold")
                text.insert("end", f"\"")
            else:
                text.insert("end", f"\"{cell_list[0]}")
                text.insert("end", f"{term}", "bold")
                text.insert("end", f"{cell_list[-1]}\"")
            text.insert("end", "\n\n")